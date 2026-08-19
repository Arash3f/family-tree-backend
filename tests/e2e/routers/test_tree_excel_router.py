from datetime import date
from io import BytesIO
from json import dumps
from uuid import UUID

import pytest
from family_tree_api_client import AuthenticatedClient
from family_tree_api_client.api.family_tree_excel.import_excel_family_trees_tree_id_excel_import_post import (  # noqa: E501
    asyncio_detailed as import_excel,
)
from family_tree_api_client.api.family_tree_excel.preview_excel_import_family_trees_tree_id_excel_import_preview_post import (  # noqa: E501
    asyncio_detailed as preview_excel_import,
)
from family_tree_api_client.api.persons.get_person_list_by_filter_family_trees_tree_id_persons_list_post import (  # noqa: E501
    asyncio_detailed as list_persons,
)
from family_tree_api_client.models.body_import_excel_family_trees_tree_id_excel_import_post import (  # noqa: E501
    BodyImportExcelFamilyTreesTreeIdExcelImportPost,
)
from family_tree_api_client.models.body_preview_excel_import_family_trees_tree_id_excel_import_preview_post import (  # noqa: E501
    BodyPreviewExcelImportFamilyTreesTreeIdExcelImportPreviewPost,
)
from family_tree_api_client.models.filter_person_request import FilterPersonRequest
from family_tree_api_client.models.paginated_response_person_model import (
    PaginatedResponsePersonModel,
)
from family_tree_api_client.models.pagination_request_params import (
    PaginationRequestParams,
)
from family_tree_api_client.models.person_filter_request_data import (
    PersonFilterRequestData,
)
from family_tree_api_client.models.person_sort_field import PersonSortField
from family_tree_api_client.models.sort_order_field import SortOrderField
from family_tree_api_client.models.sort_request_params_person_sort_field import (
    SortRequestParamsPersonSortField,
)
from family_tree_api_client.models.tree_excel_import_response import (
    TreeExcelImportResponse,
)
from family_tree_api_client.models.tree_excel_preview_response import (
    TreeExcelPreviewResponse,
)
from openpyxl import Workbook

from app.application.services.tree_excel_service import (
    MARRIAGE_HEADERS,
    PERSON_HEADERS,
)
from app.domain.entities.person import Gender, Person
from tests.e2e.auth_headers import admin_client as admin_client
from tests.helpers.uow import TreeUnitOfWork


class _PreviewBody(BodyPreviewExcelImportFamilyTreesTreeIdExcelImportPreviewPost):
    def to_multipart(self):
        return [("file", (self.file_name, self.payload, self.mime_type))]


class _ImportBody(BodyImportExcelFamilyTreesTreeIdExcelImportPost):
    def to_multipart(self):
        files = [("file", (self.file_name, self.payload, self.mime_type))]
        if self.include is not None:
            files.append(("include", (None, str(self.include).encode(), "text/plain")))
        return files


def _preview_body(content: bytes, name: str = "tree.xlsx") -> _PreviewBody:
    body = _PreviewBody(file=name)
    body.file_name = name
    body.payload = content
    body.mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return body


def _import_body(
    content: bytes, *, include: str | None = None, name: str = "tree.xlsx"
) -> _ImportBody:
    body = _ImportBody(file=name, include=include)
    body.file_name = name
    body.payload = content
    body.mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return body


async def _download_sample(
    tree_id: UUID, client: AuthenticatedClient, headers: dict[str, str] | None = None
):
    httpx_client = client.get_async_httpx_client()
    return await httpx_client.request(
        "get",
        f"/family-trees/{tree_id}/excel/sample",
        headers=headers or {},
    )


def _workbook(
    *, people: list[list[object]], marriages: list[list[object]] | None = None
) -> bytes:
    workbook = Workbook()
    workbook.active.title = "Instructions"
    persons = workbook.create_sheet("Persons")
    for index, header in enumerate(PERSON_HEADERS, start=1):
        persons.cell(row=1, column=index, value=header)
    for row_index, row in enumerate(people, start=2):
        for col_index, value in enumerate(row, start=1):
            persons.cell(row=row_index, column=col_index, value=value)
    marriages_ws = workbook.create_sheet("Marriages")
    for index, header in enumerate(MARRIAGE_HEADERS, start=1):
        marriages_ws.cell(row=1, column=index, value=header)
    for row_index, row in enumerate(marriages or [], start=2):
        for col_index, value in enumerate(row, start=1):
            marriages_ws.cell(row=row_index, column=col_index, value=value)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@pytest.mark.asyncio
async def test_sample_excel_has_no_id_column(
    tree_id, admin_client: AuthenticatedClient
):
    from openpyxl import load_workbook

    resp = await _download_sample(tree_id, admin_client)
    assert resp.status_code == 200
    workbook = load_workbook(BytesIO(resp.content), data_only=True)
    person_headers = [
        cell.value for cell in workbook["Persons"][1] if cell.value is not None
    ]
    marriage_headers = [
        cell.value for cell in workbook["Marriages"][1] if cell.value is not None
    ]
    assert "id" not in person_headers
    assert "id" not in marriage_headers


@pytest.mark.asyncio
async def test_sample_excel_uses_accept_language(
    tree_id, admin_client: AuthenticatedClient
):
    from openpyxl import load_workbook

    from app.application.services.tree_excel_service import PERSON_HEADERS_FA

    resp = await _download_sample(
        tree_id, admin_client, headers={"Accept-Language": "fa-IR"}
    )
    assert resp.status_code == 200
    assert "family-tree-sample-fa.xlsx" in resp.headers.get("content-disposition", "")
    workbook = load_workbook(BytesIO(resp.content), data_only=True)
    assert workbook["راهنما"]["A1"].value == "قالب اکسل شجره‌نامه"
    assert workbook["افراد"]["B2"].value == "علی"
    person_headers = [
        cell.value for cell in workbook["افراد"][1] if cell.value is not None
    ]
    assert person_headers == PERSON_HEADERS_FA

    resp_en = await _download_sample(tree_id, admin_client)
    assert resp_en.status_code == 200
    assert "family-tree-sample-en.xlsx" in resp_en.headers.get(
        "content-disposition", ""
    )
    workbook_en = load_workbook(BytesIO(resp_en.content), data_only=True)
    assert workbook_en["Instructions"]["A1"].value == "Family Tree Excel template"
    assert workbook_en["Persons"]["B2"].value == "Ali"


@pytest.mark.asyncio
async def test_preview_marks_people_already_in_the_tree(
    tree_id, admin_client: AuthenticatedClient, uow: TreeUnitOfWork
):
    await uow.persons.create(
        Person(
            tree_id=uow.tree_id,
            id=None,
            name="Ali",
            family_name="Karimi",
            gender=Gender.MALE,
            birth_date=date(1970, 1, 1),
        )
    )
    await uow.commit()

    content = _workbook(
        people=[
            ["P1", "Ali", "Karimi", "male", "1970-01-01"],
            ["P2", "Reza", "Karimi", "male", "1995-01-01"],
        ]
    )
    resp = await preview_excel_import(
        tree_id=tree_id, client=admin_client, body=_preview_body(content)
    )
    assert resp.status_code == 200, resp.content
    assert isinstance(resp.parsed, TreeExcelPreviewResponse)
    body = resp.parsed
    by_ref = {person.ref: person for person in body.persons}
    assert by_ref["P1"].already_exists is True
    assert "Ali Karimi" in (by_ref["P1"].existing_label or "")
    assert by_ref["P2"].already_exists is False
    assert body.valid is True


@pytest.mark.asyncio
async def test_import_skips_existing_and_only_creates_selected_people(
    tree_id, admin_client: AuthenticatedClient, uow: TreeUnitOfWork
):
    await uow.persons.create(
        Person(
            tree_id=uow.tree_id,
            id=None,
            name="Ali",
            family_name="Karimi",
            gender=Gender.MALE,
            birth_date=date(1970, 1, 1),
        )
    )
    await uow.commit()

    content = _workbook(
        people=[
            ["P1", "Ali", "Karimi", "male", "1970-01-01"],
            ["P2", "Reza", "Karimi", "male", "1995-01-01"],
            ["P3", "Sara", "Ahmadi", "female", "1996-01-01"],
        ]
    )
    resp = await import_excel(
        tree_id=tree_id,
        client=admin_client,
        body=_import_body(
            content, include=dumps({"person_refs": ["P2"], "marriage_refs": []})
        ),
    )
    assert resp.status_code == 200, resp.content
    assert isinstance(resp.parsed, TreeExcelImportResponse)
    body = resp.parsed
    assert body.persons_created == 1
    assert body.marriages_created == 0

    listed = await list_persons(
        tree_id=tree_id,
        client=admin_client,
        body=FilterPersonRequest(
            filters=PersonFilterRequestData(),
            pagination=PaginationRequestParams(offset=0, page=1, page_size=30),
            sort=SortRequestParamsPersonSortField(
                sort_by=PersonSortField.NAME, sort_order=SortOrderField.ASC
            ),
        ),
    )
    assert listed.status_code == 200, listed.content
    assert isinstance(listed.parsed, PaginatedResponsePersonModel)
    names = {item.name for item in listed.parsed.items}
    assert names == {"Ali", "Reza"}
