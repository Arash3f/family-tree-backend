from datetime import date
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock
from uuid import UUID, uuid4

import pytest
from openpyxl import Workbook

from app.application.services.tree_excel_service import MARRIAGE_HEADERS, PERSON_HEADERS
from app.application.use_cases.family_tree.import_tree_excel_use_case import (
    ImportTreeExcelUseCase,
)
from app.application.use_cases.family_tree.preview_tree_excel_use_case import (
    PreviewTreeExcelUseCase,
)
from app.domain.entities.person import Gender, Person
from app.domain.exceptions.family_tree_exceptions import TreeExcelEmptyException
from app.domain.services.marriage_rules import MarriageRulesService
from app.domain.shared.dto.pagination_dto import PaginatedResult

TREE_ID = UUID(int=11)


def _xlsx(
    *, people: list[list[object]], marriages: list[list[object]] | None = None
) -> bytes:
    workbook = Workbook()
    workbook.active.title = "Instructions"
    persons = workbook.create_sheet("Persons")
    for index, header in enumerate(PERSON_HEADERS, start=1):
        persons.cell(row=1, column=index, value=header)
    for row_index, row in enumerate(people, start=2):
        for col_index, value in enumerate(row, start=1):
            persons.cell(row=row_index, column=col_index, value=value)
    marriages_ws = workbook.create_sheet("Marriages")
    for index, header in enumerate(MARRIAGE_HEADERS, start=1):
        marriages_ws.cell(row=1, column=index, value=header)
    for row_index, row in enumerate(marriages or [], start=2):
        for col_index, value in enumerate(row, start=1):
            marriages_ws.cell(row=row_index, column=col_index, value=value)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _page(items):
    return PaginatedResult(items=items, total=len(items), page=1, page_size=100)


def _existing_ali() -> Person:
    return Person(
        id=uuid4(),
        name="Ali",
        family_name="Karimi",
        gender=Gender.MALE,
        tree_id=TREE_ID,
        birth_date=date(1970, 1, 1),
    )


@pytest.mark.asyncio
async def test_preview_marks_existing_people(mock_uow):
    existing = _existing_ali()
    mock_uow.persons.get_list_by_filter = AsyncMock(return_value=_page([existing]))
    mock_uow.marriages.get_list_by_filter = AsyncMock(return_value=_page([]))

    usecase = PreviewTreeExcelUseCase(mock_uow, MarriageRulesService())
    result = await usecase.execute(
        tree_id=TREE_ID,
        content=_xlsx(
            people=[
                ["P1", "Ali", "Karimi", "male", "1970-01-01"],
                ["P2", "Reza", "Karimi", "male", "1995-01-01"],
            ]
        ),
    )

    by_ref = {person.ref: person for person in result.persons}
    assert result.valid is True
    assert by_ref["P1"].already_exists is True
    assert existing.name in (by_ref["P1"].existing_label or "")
    assert by_ref["P2"].already_exists is False


@pytest.mark.asyncio
async def test_import_creates_only_selected_new_people(mock_uow):
    existing = _existing_ali()
    mock_uow.persons.get_list_by_filter = AsyncMock(return_value=_page([existing]))
    mock_uow.marriages.get_list_by_filter = AsyncMock(return_value=_page([]))
    created: dict[UUID, Person] = {}

    async def create_person(person: Person) -> Person:
        person.id = uuid4()
        created[person.safe_id] = person
        return person

    async def get_in_tree(person_id: UUID, tree_id: UUID) -> Person:
        return created[person_id]

    mock_uow.persons.create = AsyncMock(side_effect=create_person)
    mock_uow.persons.get_in_tree_or_raise = AsyncMock(side_effect=get_in_tree)
    mock_uow.persons.update = AsyncMock(side_effect=lambda person: person)

    usecase = ImportTreeExcelUseCase(
        mock_uow, MarriageRulesService(), sync_service=MagicMock()
    )
    result = await usecase.execute(
        tree_id=TREE_ID,
        content=_xlsx(
            people=[
                ["P1", "Ali", "Karimi", "male", "1970-01-01"],
                ["P2", "Reza", "Karimi", "male", "1995-01-01"],
                ["P3", "Sara", "Ahmadi", "female", "1996-01-01"],
            ]
        ),
        person_refs={"P2"},
        marriage_refs=set(),
    )

    assert result.persons_created == 1
    assert result.marriages_created == 0
    created_names = {person.name for person in created.values()}
    assert created_names == {"Reza"}
    mock_uow.persons.create.assert_awaited_once()


@pytest.mark.asyncio
async def test_import_without_selection_skips_existing_people(mock_uow):
    existing = _existing_ali()
    mock_uow.persons.get_list_by_filter = AsyncMock(return_value=_page([existing]))
    mock_uow.marriages.get_list_by_filter = AsyncMock(return_value=_page([]))
    created: dict[UUID, Person] = {}

    async def create_person(person: Person) -> Person:
        person.id = uuid4()
        created[person.safe_id] = person
        return person

    mock_uow.persons.create = AsyncMock(side_effect=create_person)
    mock_uow.persons.get_in_tree_or_raise = AsyncMock(
        side_effect=lambda person_id, tree_id: created[person_id]
    )
    mock_uow.persons.update = AsyncMock(side_effect=lambda person: person)

    usecase = ImportTreeExcelUseCase(
        mock_uow, MarriageRulesService(), sync_service=MagicMock()
    )
    result = await usecase.execute(
        tree_id=TREE_ID,
        content=_xlsx(
            people=[
                ["P1", "Ali", "Karimi", "male", "1970-01-01"],
                ["P2", "Reza", "Karimi", "male", "1995-01-01"],
            ]
        ),
    )

    assert result.persons_created == 1
    assert {person.name for person in created.values()} == {"Reza"}


@pytest.mark.asyncio
async def test_import_selected_existing_only_is_empty(mock_uow):
    existing = _existing_ali()
    mock_uow.persons.get_list_by_filter = AsyncMock(return_value=_page([existing]))
    mock_uow.marriages.get_list_by_filter = AsyncMock(return_value=_page([]))

    usecase = ImportTreeExcelUseCase(
        mock_uow, MarriageRulesService(), sync_service=MagicMock()
    )
    with pytest.raises(TreeExcelEmptyException):
        await usecase.execute(
            tree_id=TREE_ID,
            content=_xlsx(people=[["P1", "Ali", "Karimi", "male", "1970-01-01"]]),
            person_refs={"P1"},
            marriage_refs=set(),
        )
    mock_uow.persons.create.assert_not_awaited()
