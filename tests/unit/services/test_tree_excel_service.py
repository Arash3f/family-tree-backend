from datetime import date
from io import BytesIO
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from app.application.services.tree_excel_service import (
    MARRIAGE_HEADERS,
    MARRIAGE_HEADERS_FA,
    PERSON_HEADERS,
    PERSON_HEADERS_FA,
    ExcelMarriageRow,
    ExcelPersonRow,
    ParsedTreeExcel,
    build_export_workbook,
    build_sample_workbook,
    match_tree_excel,
    parse_tree_excel,
)
from app.domain.entities.marriage import Marriage
from app.domain.entities.person import Gender, Person
from app.presentation.utils.date_convert import gregorian_to_jalali


def test_excel_headers_do_not_include_id():
    assert "id" not in PERSON_HEADERS
    assert "id" not in MARRIAGE_HEADERS
    assert "id" not in PERSON_HEADERS_FA
    assert "id" not in MARRIAGE_HEADERS_FA


def test_sample_workbook_has_no_id_column():
    workbook = load_workbook(BytesIO(build_sample_workbook()), data_only=True)
    person_headers = [
        cell.value for cell in workbook["Persons"][1] if cell.value is not None
    ]
    marriage_headers = [
        cell.value for cell in workbook["Marriages"][1] if cell.value is not None
    ]
    assert person_headers == PERSON_HEADERS
    assert marriage_headers == MARRIAGE_HEADERS


def test_sample_workbook_fa_has_persian_sheets_headers_and_data():
    workbook = load_workbook(BytesIO(build_sample_workbook(lang="fa")), data_only=True)
    assert workbook["راهنما"]["A1"].value == "قالب اکسل شجره‌نامه"
    assert workbook["افراد"]["B2"].value == "علی"
    assert workbook["افراد"]["C2"].value == "کریمی"
    assert workbook["افراد"]["D2"].value == "مرد"
    person_headers = [
        cell.value for cell in workbook["افراد"][1] if cell.value is not None
    ]
    marriage_headers = [
        cell.value for cell in workbook["ازدواج‌ها"][1] if cell.value is not None
    ]
    assert person_headers == PERSON_HEADERS_FA
    assert marriage_headers == MARRIAGE_HEADERS_FA
    assert workbook["ازدواج‌ها"]["D2"].value == "1334/03/10"
    parsed = parse_tree_excel(build_sample_workbook(lang="fa"))
    assert [person.ref for person in parsed.persons] == ["P1", "P2", "P3", "P4", "P5"]
    assert parsed.persons[0].gender == Gender.MALE
    assert parsed.persons[0].birth_date == date(1951, 3, 22)


def test_sample_workbook_en_has_english_instructions_and_data():
    workbook = load_workbook(BytesIO(build_sample_workbook(lang="en")), data_only=True)
    assert workbook["Instructions"]["A1"].value == "Family Tree Excel template"
    assert workbook["Persons"]["B2"].value == "Ali"
    assert workbook["Persons"]["C2"].value == "Karimi"


def test_parse_tree_excel_accepts_workbook_without_id():
    parsed = parse_tree_excel(build_sample_workbook())
    assert [person.ref for person in parsed.persons] == ["P1", "P2", "P3", "P4", "P5"]
    assert [marriage.ref for marriage in parsed.marriages] == ["M1", "M2"]


def test_export_workbook_has_no_id_column():
    tree_id = uuid4()
    person = Person(
        id=uuid4(),
        name="Ali",
        family_name="Karimi",
        gender=Gender.MALE,
        tree_id=tree_id,
        birth_date=date(1951, 3, 22),
    )
    content = build_export_workbook(persons=[person], marriages=[])
    workbook = load_workbook(BytesIO(content), data_only=True)
    headers = [cell.value for cell in workbook["Persons"][1] if cell.value is not None]
    assert "id" not in headers
    assert headers == PERSON_HEADERS
    assert workbook["Persons"]["E2"].value == "1951-03-22"


def test_export_workbook_fa_uses_persian_headers_and_jalali_dates():
    tree_id = uuid4()
    person = Person(
        id=uuid4(),
        name="علی",
        family_name="کریمی",
        gender=Gender.MALE,
        tree_id=tree_id,
        birth_date=date(1951, 3, 22),
    )
    content = build_export_workbook(persons=[person], marriages=[], lang="fa")
    workbook = load_workbook(BytesIO(content), data_only=True)
    headers = [cell.value for cell in workbook["افراد"][1] if cell.value is not None]
    assert headers == PERSON_HEADERS_FA
    assert workbook["افراد"]["D2"].value == "مرد"
    assert workbook["افراد"]["E2"].value == gregorian_to_jalali(date(1951, 3, 22))
    parsed = parse_tree_excel(content)
    assert parsed.persons[0].birth_date == date(1951, 3, 22)
    assert parsed.persons[0].gender == Gender.MALE


def test_match_tree_excel_detects_existing_and_in_file_duplicates():
    tree_id = uuid4()
    existing_id = uuid4()
    existing = Person(
        id=existing_id,
        name="Ali",
        family_name="Karimi",
        gender=Gender.MALE,
        tree_id=tree_id,
        birth_date=date(1951, 3, 22),
    )
    parsed = ParsedTreeExcel(
        persons=[
            ExcelPersonRow(
                ref="P1",
                name="Ali",
                family_name="Karimi",
                gender=Gender.MALE,
                birth_date=date(1951, 3, 22),
                row_number=2,
            ),
            ExcelPersonRow(
                ref="P2",
                name="Ali",
                family_name="Karimi",
                gender=Gender.MALE,
                birth_date=date(1951, 3, 22),
                row_number=3,
            ),
            ExcelPersonRow(
                ref="P3",
                name="Reza",
                gender=Gender.MALE,
                row_number=4,
            ),
        ]
    )

    match = match_tree_excel(parsed, [existing], [])

    assert match.person_existing_id["P1"] == existing_id
    assert match.person_existing_id["P2"] == existing_id
    assert match.person_duplicate_of["P2"] == "P1"
    assert "P3" not in match.person_existing_id
    assert "P3" not in match.person_duplicate_of


def test_match_tree_excel_detects_existing_marriage():
    tree_id = uuid4()
    husband_id = uuid4()
    wife_id = uuid4()
    marriage_id = uuid4()
    husband = Person(
        id=husband_id,
        name="Ali",
        gender=Gender.MALE,
        tree_id=tree_id,
        birth_date=date(1970, 1, 1),
    )
    wife = Person(
        id=wife_id,
        name="Zahra",
        gender=Gender.FEMALE,
        tree_id=tree_id,
        birth_date=date(1972, 1, 1),
    )
    marriage = Marriage(
        id=marriage_id,
        tree_id=tree_id,
        spouse_a_id=husband_id,
        spouse_b_id=wife_id,
        married_at=date(1995, 6, 1),
    )

    parsed = ParsedTreeExcel(
        persons=[
            ExcelPersonRow(
                ref="P1",
                name="Ali",
                gender=Gender.MALE,
                birth_date=date(1970, 1, 1),
                row_number=2,
            ),
            ExcelPersonRow(
                ref="P2",
                name="Zahra",
                gender=Gender.FEMALE,
                birth_date=date(1972, 1, 1),
                row_number=3,
            ),
        ],
        marriages=[
            ExcelMarriageRow(
                ref="M1",
                spouse_a_ref="P1",
                spouse_b_ref="P2",
                married_at=date(1995, 6, 1),
                row_number=2,
            )
        ],
    )

    match = match_tree_excel(parsed, [husband, wife], [marriage])

    assert match.marriage_existing_id["M1"] == marriage_id


def test_parse_ignores_legacy_id_column_if_present():
    workbook = Workbook()
    workbook.active.title = "Instructions"
    persons = workbook.create_sheet("Persons")
    headers = [*PERSON_HEADERS, "id"]
    for index, header in enumerate(headers, start=1):
        persons.cell(row=1, column=index, value=header)
    persons.cell(row=2, column=1, value="P1")
    persons.cell(row=2, column=2, value="Ali")
    persons.cell(row=2, column=4, value="male")
    persons.cell(row=2, column=len(headers), value=str(uuid4()))
    marriages = workbook.create_sheet("Marriages")
    for index, header in enumerate(MARRIAGE_HEADERS, start=1):
        marriages.cell(row=1, column=index, value=header)
    buffer = BytesIO()
    workbook.save(buffer)

    parsed = parse_tree_excel(buffer.getvalue())
    assert [person.ref for person in parsed.persons] == ["P1"]
    assert parsed.persons[0].name == "Ali"
