from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.domain.entities.marriage import Marriage
from app.domain.entities.person import (
    Gender,
    ParentRelationshipType,
    Person,
)
from app.domain.exceptions.family_tree_exceptions import TreeExcelInvalidException
from app.presentation.utils.date_convert import gregorian_to_jalali, parse_user_date

PERSONS_SHEET = "Persons"
MARRIAGES_SHEET = "Marriages"
INSTRUCTIONS_SHEET = "Instructions"

PERSONS_SHEET_FA = "افراد"
MARRIAGES_SHEET_FA = "ازدواج‌ها"
INSTRUCTIONS_SHEET_FA = "راهنما"

PERSONS_SHEET_ALIASES = (PERSONS_SHEET, PERSONS_SHEET_FA)
MARRIAGES_SHEET_ALIASES = (MARRIAGES_SHEET, MARRIAGES_SHEET_FA)

PERSON_HEADERS = [
    "ref",
    "name",
    "family_name",
    "gender",
    "birth_date",
    "death_date",
    "birth_place",
    "death_place",
    "notes",
    "parent1_ref",
    "parent1_type",
    "parent2_ref",
    "parent2_type",
    "marriage_ref",
]

MARRIAGE_HEADERS = [
    "ref",
    "spouse_a_ref",
    "spouse_b_ref",
    "married_at",
    "divorced_at",
]

PERSON_HEADERS_FA = [
    "شناسه",
    "نام",
    "نام خانوادگی",
    "جنسیت",
    "تاریخ تولد",
    "تاریخ فوت",
    "محل تولد",
    "محل فوت",
    "یادداشت",
    "شناسه والد ۱",
    "نوع والد ۱",
    "شناسه والد ۲",
    "نوع والد ۲",
    "شناسه ازدواج",
]

MARRIAGE_HEADERS_FA = [
    "شناسه",
    "شناسه همسر ۱",
    "شناسه همسر ۲",
    "تاریخ ازدواج",
    "تاریخ طلاق",
]

_PERSON_HEADER_ALIASES: dict[str, frozenset[str]] = {
    "ref": frozenset({"ref", "شناسه"}),
    "name": frozenset({"name", "نام"}),
    "family_name": frozenset({"family_name", "نام خانوادگی"}),
    "gender": frozenset({"gender", "جنسیت"}),
    "birth_date": frozenset({"birth_date", "تاریخ تولد", "تولد"}),
    "death_date": frozenset({"death_date", "تاریخ فوت", "فوت"}),
    "birth_place": frozenset({"birth_place", "محل تولد"}),
    "death_place": frozenset({"death_place", "محل فوت"}),
    "notes": frozenset({"notes", "یادداشت"}),
    "parent1_ref": frozenset({"parent1_ref", "شناسه والد ۱", "شناسه والد 1"}),
    "parent1_type": frozenset({"parent1_type", "نوع والد ۱", "نوع والد 1"}),
    "parent2_ref": frozenset({"parent2_ref", "شناسه والد ۲", "شناسه والد 2"}),
    "parent2_type": frozenset({"parent2_type", "نوع والد ۲", "نوع والد 2"}),
    "marriage_ref": frozenset({"marriage_ref", "شناسه ازدواج", "ازدواج مبدأ"}),
}

_MARRIAGE_HEADER_ALIASES: dict[str, frozenset[str]] = {
    "ref": frozenset({"ref", "شناسه"}),
    "spouse_a_ref": frozenset(
        {"spouse_a_ref", "شناسه همسر ۱", "شناسه همسر 1", "همسر ۱", "همسر 1"}
    ),
    "spouse_b_ref": frozenset(
        {"spouse_b_ref", "شناسه همسر ۲", "شناسه همسر 2", "همسر ۲", "همسر 2"}
    ),
    "married_at": frozenset({"married_at", "تاریخ ازدواج"}),
    "divorced_at": frozenset({"divorced_at", "تاریخ طلاق"}),
}

_GENDER_ALIASES: dict[str, frozenset[str]] = {
    Gender.MALE.value: frozenset({"male", "مرد"}),
    Gender.FEMALE.value: frozenset({"female", "زن"}),
}

_REL_TYPE_ALIASES: dict[str, frozenset[str]] = {
    ParentRelationshipType.BIOLOGICAL.value: frozenset({"biological", "بیولوژیک"}),
    ParentRelationshipType.ADOPTIVE.value: frozenset({"adoptive", "فرزندخواندگی"}),
    ParentRelationshipType.STEP.value: frozenset({"step", "ناتنی"}),
}

HEADER_FILL = PatternFill("solid", fgColor="0D6E67")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SAMPLE_FILL = PatternFill("solid", fgColor="E8F5F3")


def _normalize_locale(lang: str | None) -> str:
    return "fa" if (lang or "").lower().startswith("fa") else "en"


def _sheet_titles(lang: str) -> tuple[str, str, str]:
    if lang == "fa":
        return PERSONS_SHEET_FA, MARRIAGES_SHEET_FA, INSTRUCTIONS_SHEET_FA
    return PERSONS_SHEET, MARRIAGES_SHEET, INSTRUCTIONS_SHEET


def _person_headers(lang: str) -> list[str]:
    return PERSON_HEADERS_FA if lang == "fa" else PERSON_HEADERS


def _marriage_headers(lang: str) -> list[str]:
    return MARRIAGE_HEADERS_FA if lang == "fa" else MARRIAGE_HEADERS


def format_excel_date(value: date | None, *, lang: str) -> str:
    if value is None:
        return ""
    if lang == "fa":
        return gregorian_to_jalali(value)
    return value.isoformat()


def _gender_label(gender: Gender, *, lang: str) -> str:
    if lang == "fa":
        return "مرد" if gender == Gender.MALE else "زن"
    return gender.value


def _rel_type_label(rel_type: ParentRelationshipType, *, lang: str) -> str:
    if lang != "fa":
        return rel_type.value
    return {
        ParentRelationshipType.BIOLOGICAL: "بیولوژیک",
        ParentRelationshipType.ADOPTIVE: "فرزندخواندگی",
        ParentRelationshipType.STEP: "ناتنی",
    }[rel_type]


def _alias_lookup(aliases: dict[str, frozenset[str]]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for canonical, names in aliases.items():
        for name in names:
            lookup[name.casefold()] = canonical
    return lookup


_PERSON_HEADER_LOOKUP = _alias_lookup(_PERSON_HEADER_ALIASES)
_MARRIAGE_HEADER_LOOKUP = _alias_lookup(_MARRIAGE_HEADER_ALIASES)
_GENDER_LOOKUP = _alias_lookup(_GENDER_ALIASES)
_REL_TYPE_LOOKUP = _alias_lookup(_REL_TYPE_ALIASES)


@dataclass
class ExcelPersonRow:
    ref: str
    name: str
    gender: Gender
    family_name: str | None = None
    birth_date: date | None = None
    death_date: date | None = None
    birth_place: str | None = None
    death_place: str | None = None
    notes: str | None = None
    parent1_ref: str | None = None
    parent1_type: ParentRelationshipType = ParentRelationshipType.BIOLOGICAL
    parent2_ref: str | None = None
    parent2_type: ParentRelationshipType = ParentRelationshipType.BIOLOGICAL
    marriage_ref: str | None = None
    row_number: int = 0


@dataclass
class ExcelMarriageRow:
    ref: str
    spouse_a_ref: str
    spouse_b_ref: str
    married_at: date
    divorced_at: date | None = None
    row_number: int = 0


@dataclass
class ParsedTreeExcel:
    persons: list[ExcelPersonRow] = field(default_factory=list)
    marriages: list[ExcelMarriageRow] = field(default_factory=list)


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def parse_excel_date(value: Any, *, field_name: str, row_number: int) -> date | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    try:
        parsed = parse_user_date(text)
    except Exception as exc:
        raise TreeExcelInvalidException(
            detail=[
                f"Row {row_number}: invalid {field_name} '{text}'. "
                "Use Jalali YYYY/MM/DD or Gregorian YYYY-MM-DD."
            ]
        ) from exc
    if parsed is None:
        raise TreeExcelInvalidException(
            detail=[
                f"Row {row_number}: invalid {field_name} '{text}'. "
                "Use Jalali YYYY/MM/DD or Gregorian YYYY-MM-DD."
            ]
        )
    return parsed


def _parse_gender(value: Any, *, row_number: int) -> Gender:
    text = (_cell_str(value) or "").casefold()
    canonical = _GENDER_LOOKUP.get(text)
    if canonical is None:
        raise TreeExcelInvalidException(
            detail=[f"Persons row {row_number}: gender must be male/female or مرد/زن"]
        )
    return Gender(canonical)


def _parse_rel_type(value: Any, *, row_number: int) -> ParentRelationshipType:
    raw = _cell_str(value)
    if not raw:
        return ParentRelationshipType.BIOLOGICAL
    canonical = _REL_TYPE_LOOKUP.get(raw.casefold())
    if canonical is None:
        raise TreeExcelInvalidException(
            detail=[
                f"Persons row {row_number}: parent type must be "
                "biological/adoptive/step or بیولوژیک/فرزندخواندگی/ناتنی"
            ]
        )
    return ParentRelationshipType(canonical)


def _style_header(ws, headers: list[str]) -> None:
    for index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for index in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(index)].width = 16


def _add_person_validations(ws, *, lang: str = "en") -> None:
    if lang == "fa":
        gender_list = '"مرد,زن"'
        gender_error = "از مرد یا زن استفاده کنید"
        type_list = '"بیولوژیک,فرزندخواندگی,ناتنی"'
        type_error = "از بیولوژیک، فرزندخواندگی یا ناتنی استفاده کنید"
    else:
        gender_list = '"male,female"'
        gender_error = "Use male or female"
        type_list = '"biological,adoptive,step"'
        type_error = "Use biological, adoptive, or step"

    gender_dv = DataValidation(
        type="list",
        formula1=gender_list,
        allow_blank=True,
        showErrorMessage=True,
    )
    gender_dv.error = gender_error
    gender_dv.add("D2:D1000")
    ws.add_data_validation(gender_dv)

    type_dv = DataValidation(
        type="list",
        formula1=type_list,
        allow_blank=True,
        showErrorMessage=True,
    )
    type_dv.error = type_error
    type_dv.add("K2:K1000")
    type_dv.add("M2:M1000")
    ws.add_data_validation(type_dv)


_SAMPLE_INSTRUCTIONS: dict[str, list[str]] = {
    "en": [
        "Family Tree Excel template",
        "",
        "1) Fill the Persons sheet. Each person needs a unique ref (e.g. P1, P2).",
        "2) Fill the Marriages sheet. Use person refs in spouse_a_ref / spouse_b_ref.",
        "3) Link children with parent1_ref / parent2_ref and optional marriage_ref.",
        "4) gender: male | female",
        "5) parent types: biological | adoptive | step",
        "6) Dates: Gregorian YYYY-MM-DD (Jalali YYYY/MM/DD is also accepted)",
        "7) Marriage dates use the same formats as person dates.",
        "8) Do not include an id column. Matching uses name, family name, gender, and birth date.",
        "9) Import lets you pick rows; people/marriages already in the tree are skipped.",
    ],
    "fa": [
        "قالب اکسل شجره‌نامه",
        "",
        "۱) برگه «افراد» را پر کنید. هر نفر یک شناسه یکتا داشته باشد (مثل P1).",
        "۲) برگه «ازدواج‌ها» را با شناسه افراد در «شناسه همسر ۱» و «شناسه همسر ۲» پر کنید.",
        "۳) برای فرزند «شناسه والد ۱ / ۲» و در صورت نیاز «شناسه ازدواج» را بگذارید.",
        "۴) جنسیت: مرد | زن",
        "۵) نوع والد: بیولوژیک | فرزندخواندگی | ناتنی",
        "۶) تاریخ‌ها شمسی با قالب YYYY/MM/DD (میلادی YYYY-MM-DD هم پذیرفته می‌شود)",
        "۷) تاریخ ازدواج هم با همان قالب‌ها وارد شود.",
        "۸) ستون id لازم نیست. افراد تکراری با نام، نام خانوادگی، جنسیت و تاریخ تولد تشخیص داده می‌شوند.",
        "۹) هنگام ورود می‌توانید ردیف‌ها را انتخاب کنید؛ افراد/ازدواج‌های موجود در درخت نادیده گرفته می‌شوند.",
        "۱۰) فایل‌های انگلیسی (Persons / Marriages و ستون‌های انگلیسی) هم قابل ورود هستند.",
    ],
}

_EXPORT_INSTRUCTIONS: dict[str, list[str]] = {
    "en": [
        "Exported family tree data",
        "All dates are Gregorian YYYY-MM-DD. Jalali YYYY/MM/DD is also accepted on import.",
        "Re-import matches existing people by name, family name, gender, and birth date.",
    ],
    "fa": [
        "داده‌های خروجی شجره‌نامه",
        "همه تاریخ‌ها شمسی با قالب YYYY/MM/DD هستند. میلادی YYYY-MM-DD هم هنگام ورود پذیرفته می‌شود.",
        "ورود مجدد افراد موجود را با نام، نام خانوادگی، جنسیت و تاریخ تولد تطبیق می‌دهد.",
    ],
}

_SAMPLE_PEOPLE: dict[str, list[list[str]]] = {
    "en": [
        [
            "P1",
            "Ali",
            "Karimi",
            "male",
            "1330/01/01",
            "",
            "Tehran",
            "",
            "Grandfather",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "P2",
            "Zahra",
            "Karimi",
            "female",
            "1335/05/10",
            "",
            "Tehran",
            "",
            "Grandmother",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "P3",
            "Reza",
            "Karimi",
            "male",
            "1358/03/20",
            "",
            "Tehran",
            "",
            "Father",
            "P1",
            "biological",
            "P2",
            "biological",
            "M1",
        ],
        [
            "P4",
            "Sara",
            "Ahmadi",
            "female",
            "1360/07/01",
            "",
            "Isfahan",
            "",
            "Mother",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "P5",
            "Arash",
            "Karimi",
            "male",
            "1379/09/01",
            "",
            "Tehran",
            "",
            "Child",
            "P3",
            "biological",
            "P4",
            "biological",
            "M2",
        ],
    ],
    "fa": [
        [
            "P1",
            "علی",
            "کریمی",
            "مرد",
            "1330/01/01",
            "",
            "تهران",
            "",
            "پدربزرگ",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "P2",
            "زهرا",
            "کریمی",
            "زن",
            "1335/05/10",
            "",
            "تهران",
            "",
            "مادربزرگ",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "P3",
            "رضا",
            "کریمی",
            "مرد",
            "1358/03/20",
            "",
            "تهران",
            "",
            "پدر",
            "P1",
            "بیولوژیک",
            "P2",
            "بیولوژیک",
            "M1",
        ],
        [
            "P4",
            "سارا",
            "احمدی",
            "زن",
            "1360/07/01",
            "",
            "اصفهان",
            "",
            "مادر",
            "",
            "",
            "",
            "",
            "",
        ],
        [
            "P5",
            "آرش",
            "کریمی",
            "مرد",
            "1379/09/01",
            "",
            "تهران",
            "",
            "فرزند",
            "P3",
            "بیولوژیک",
            "P4",
            "بیولوژیک",
            "M2",
        ],
    ],
}


def build_sample_workbook(*, lang: str = "en") -> bytes:
    locale = _normalize_locale(lang)
    persons_title, marriages_title, instructions_title = _sheet_titles(locale)
    wb = Workbook()

    instructions = wb.active
    instructions.title = instructions_title
    lines = _SAMPLE_INSTRUCTIONS[locale]
    for index, line in enumerate(lines, start=1):
        instructions.cell(row=index, column=1, value=line)
    instructions.column_dimensions["A"].width = 100

    persons = wb.create_sheet(persons_title)
    _style_header(persons, _person_headers(locale))
    _add_person_validations(persons, lang=locale)
    sample_people = _SAMPLE_PEOPLE[locale]
    for row_index, row in enumerate(sample_people, start=2):
        for col_index, value in enumerate(row, start=1):
            cell = persons.cell(row=row_index, column=col_index, value=value)
            cell.fill = SAMPLE_FILL

    marriages = wb.create_sheet(marriages_title)
    _style_header(marriages, _marriage_headers(locale))
    if locale == "fa":
        sample_marriages = [
            ["M1", "P1", "P2", "1334/03/10", ""],
            ["M2", "P3", "P4", "1364/01/23", ""],
        ]
    else:
        sample_marriages = [
            ["M1", "P1", "P2", "1955-06-01", ""],
            ["M2", "P3", "P4", "1985-04-12", ""],
        ]
    for row_index, row in enumerate(sample_marriages, start=2):
        for col_index, value in enumerate(row, start=1):
            cell = marriages.cell(row=row_index, column=col_index, value=value)
            cell.fill = SAMPLE_FILL

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_export_workbook(
    *,
    persons: list[Person],
    marriages: list[Marriage],
    lang: str = "en",
) -> bytes:
    locale = _normalize_locale(lang)
    persons_title, marriages_title, instructions_title = _sheet_titles(locale)
    wb = Workbook()
    instructions = wb.active
    instructions.title = instructions_title
    for index, line in enumerate(_EXPORT_INSTRUCTIONS[locale], start=1):
        instructions.cell(row=index, column=1, value=line)
    instructions.column_dimensions["A"].width = 90

    person_ref_by_id = {
        person.safe_id: f"P{index}" for index, person in enumerate(persons, start=1)
    }
    marriage_ref_by_id = {
        marriage.safe_id: f"M{index}"
        for index, marriage in enumerate(marriages, start=1)
    }

    persons_ws = wb.create_sheet(persons_title)
    _style_header(persons_ws, _person_headers(locale))
    _add_person_validations(persons_ws, lang=locale)

    for row_index, person in enumerate(persons, start=2):
        parents = list(person.parents)
        p1 = parents[0] if len(parents) > 0 else None
        p2 = parents[1] if len(parents) > 1 else None
        values = [
            person_ref_by_id[person.safe_id],
            person.name,
            person.family_name or "",
            _gender_label(person.gender, lang=locale),
            format_excel_date(person.birth_date, lang=locale),
            format_excel_date(person.death_date, lang=locale),
            person.birth_place or "",
            person.death_place or "",
            person.notes or "",
            person_ref_by_id.get(p1.parent_id, "") if p1 else "",
            _rel_type_label(p1.relationship_type, lang=locale) if p1 else "",
            person_ref_by_id.get(p2.parent_id, "") if p2 else "",
            _rel_type_label(p2.relationship_type, lang=locale) if p2 else "",
            marriage_ref_by_id.get(person.marriage_id, "")
            if person.marriage_id
            else "",
        ]
        for col_index, value in enumerate(values, start=1):
            persons_ws.cell(row=row_index, column=col_index, value=value)

    marriages_ws = wb.create_sheet(marriages_title)
    _style_header(marriages_ws, _marriage_headers(locale))
    for row_index, marriage in enumerate(marriages, start=2):
        values = [
            marriage_ref_by_id[marriage.safe_id],
            person_ref_by_id.get(marriage.spouse_a_id, str(marriage.spouse_a_id)),
            person_ref_by_id.get(marriage.spouse_b_id, str(marriage.spouse_b_id)),
            format_excel_date(marriage.married_at, lang=locale),
            format_excel_date(marriage.divorced_at, lang=locale),
        ]
        for col_index, value in enumerate(values, start=1):
            marriages_ws.cell(row=row_index, column=col_index, value=value)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _resolve_sheet(wb, aliases: tuple[str, ...], *, label: str):
    by_key = {name.strip().casefold(): name for name in wb.sheetnames}
    for alias in aliases:
        found = by_key.get(alias.strip().casefold())
        if found is not None:
            return wb[found]
    raise TreeExcelInvalidException(
        detail=[f"Missing required sheet '{label}' ({' / '.join(aliases)})"]
    )


def _header_map(ws, aliases: dict[str, frozenset[str]]) -> dict[str, int]:
    lookup = _alias_lookup(aliases)
    mapping: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=col).value
        if raw is None:
            continue
        key = str(raw).strip().casefold()
        canonical = lookup.get(key, key)
        mapping[canonical] = col
    return mapping


def _require_headers(mapping: dict[str, int], required: list[str], sheet: str) -> None:
    missing = [name for name in required if name not in mapping]
    if missing:
        raise TreeExcelInvalidException(
            detail=[f"{sheet} sheet missing columns: {', '.join(missing)}"]
        )


def parse_tree_excel(content: bytes) -> ParsedTreeExcel:
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise TreeExcelInvalidException(
            detail=["Could not read Excel file. Upload a valid .xlsx workbook."]
        ) from exc

    persons_ws = _resolve_sheet(wb, PERSONS_SHEET_ALIASES, label=PERSONS_SHEET)
    marriages_ws = _resolve_sheet(wb, MARRIAGES_SHEET_ALIASES, label=MARRIAGES_SHEET)
    person_cols = _header_map(persons_ws, _PERSON_HEADER_ALIASES)
    marriage_cols = _header_map(marriages_ws, _MARRIAGE_HEADER_ALIASES)

    _require_headers(
        person_cols,
        ["ref", "name", "gender"],
        PERSONS_SHEET,
    )
    _require_headers(
        marriage_cols,
        ["ref", "spouse_a_ref", "spouse_b_ref", "married_at"],
        MARRIAGES_SHEET,
    )

    persons: list[ExcelPersonRow] = []
    seen_person_refs: set[str] = set()
    for row_number in range(2, persons_ws.max_row + 1):

        def get(name: str) -> Any:
            col = person_cols.get(name)
            if col is None:
                return None
            return persons_ws.cell(row=row_number, column=col).value

        ref = _cell_str(get("ref"))
        name = _cell_str(get("name"))
        gender_raw = get("gender")
        if not ref and not name and not _cell_str(gender_raw):
            continue
        if not ref or not name:
            raise TreeExcelInvalidException(
                detail=[f"Persons row {row_number}: ref and name are required"]
            )
        if ref in seen_person_refs:
            raise TreeExcelInvalidException(
                detail=[f"Persons row {row_number}: duplicate ref '{ref}'"]
            )
        seen_person_refs.add(ref)

        persons.append(
            ExcelPersonRow(
                ref=ref,
                name=name,
                gender=_parse_gender(gender_raw, row_number=row_number),
                family_name=_cell_str(get("family_name")),
                birth_date=parse_excel_date(
                    get("birth_date"), field_name="birth_date", row_number=row_number
                ),
                death_date=parse_excel_date(
                    get("death_date"), field_name="death_date", row_number=row_number
                ),
                birth_place=_cell_str(get("birth_place")),
                death_place=_cell_str(get("death_place")),
                notes=_cell_str(get("notes")),
                parent1_ref=_cell_str(get("parent1_ref")),
                parent1_type=_parse_rel_type(
                    get("parent1_type"), row_number=row_number
                ),
                parent2_ref=_cell_str(get("parent2_ref")),
                parent2_type=_parse_rel_type(
                    get("parent2_type"), row_number=row_number
                ),
                marriage_ref=_cell_str(get("marriage_ref")),
                row_number=row_number,
            )
        )

    marriages: list[ExcelMarriageRow] = []
    seen_marriage_refs: set[str] = set()
    for row_number in range(2, marriages_ws.max_row + 1):

        def get(name: str) -> Any:
            col = marriage_cols.get(name)
            if col is None:
                return None
            return marriages_ws.cell(row=row_number, column=col).value

        ref = _cell_str(get("ref"))
        spouse_a = _cell_str(get("spouse_a_ref"))
        spouse_b = _cell_str(get("spouse_b_ref"))
        married_raw = get("married_at")
        if not ref and not spouse_a and not spouse_b and not _cell_str(married_raw):
            continue
        if not ref or not spouse_a or not spouse_b:
            raise TreeExcelInvalidException(
                detail=[
                    f"Marriages row {row_number}: ref, spouse_a_ref and spouse_b_ref "
                    "are required"
                ]
            )
        if ref in seen_marriage_refs:
            raise TreeExcelInvalidException(
                detail=[f"Marriages row {row_number}: duplicate ref '{ref}'"]
            )
        seen_marriage_refs.add(ref)
        married_at = parse_excel_date(
            married_raw, field_name="married_at", row_number=row_number
        )
        if married_at is None:
            raise TreeExcelInvalidException(
                detail=[f"Marriages row {row_number}: married_at is required"]
            )
        marriages.append(
            ExcelMarriageRow(
                ref=ref,
                spouse_a_ref=spouse_a,
                spouse_b_ref=spouse_b,
                married_at=married_at,
                divorced_at=parse_excel_date(
                    get("divorced_at"),
                    field_name="divorced_at",
                    row_number=row_number,
                ),
                row_number=row_number,
            )
        )

    return ParsedTreeExcel(persons=persons, marriages=marriages)


PersonIdentityKey = tuple[str, str, str, date | None]
MarriageFileKey = tuple[frozenset[str], date]
MarriageExistingKey = tuple[frozenset[UUID], date]


def person_identity_key(
    name: str,
    family_name: str | None,
    gender: Gender | str,
    birth_date: date | None,
) -> PersonIdentityKey:
    gender_value = gender.value if isinstance(gender, Gender) else str(gender)
    return (
        name.strip().casefold(),
        (family_name or "").strip().casefold(),
        gender_value.strip().lower(),
        birth_date,
    )


def person_display_label(person: Person) -> str:
    parts = [person.name]
    if person.family_name:
        parts.append(person.family_name)
    label = " ".join(parts)
    if person.birth_date:
        return f"{label} ({person.birth_date.isoformat()})"
    return label


def canonical_excel_ref(ref: str, duplicate_of: dict[str, str]) -> str:
    seen: set[str] = set()
    current = ref
    while current in duplicate_of and current not in seen:
        seen.add(current)
        current = duplicate_of[current]
    return current


@dataclass
class TreeExcelMatch:
    person_existing_id: dict[str, UUID] = field(default_factory=dict)
    person_existing_label: dict[str, str] = field(default_factory=dict)
    person_duplicate_of: dict[str, str] = field(default_factory=dict)
    marriage_existing_id: dict[str, UUID] = field(default_factory=dict)
    marriage_duplicate_of: dict[str, str] = field(default_factory=dict)

    def person_already_in_tree(self, ref: str) -> bool:
        return ref in self.person_existing_id

    def marriage_already_in_tree(self, ref: str) -> bool:
        return ref in self.marriage_existing_id


def match_tree_excel(
    parsed: ParsedTreeExcel,
    existing_persons: list[Person],
    existing_marriages: list[Marriage],
) -> TreeExcelMatch:
    existing_by_key: dict[PersonIdentityKey, Person] = {}
    for person in existing_persons:
        key = person_identity_key(
            person.name, person.family_name, person.gender, person.birth_date
        )
        existing_by_key.setdefault(key, person)

    match = TreeExcelMatch()
    file_key_to_ref: dict[PersonIdentityKey, str] = {}
    for row in parsed.persons:
        key = person_identity_key(row.name, row.family_name, row.gender, row.birth_date)
        if key in file_key_to_ref:
            match.person_duplicate_of[row.ref] = file_key_to_ref[key]
        else:
            file_key_to_ref[key] = row.ref
        existing = existing_by_key.get(key)
        if existing is not None:
            match.person_existing_id[row.ref] = existing.safe_id
            match.person_existing_label[row.ref] = person_display_label(existing)

    existing_marriage_by_key: dict[MarriageExistingKey, Marriage] = {}
    for marriage in existing_marriages:
        marriage_key = (
            frozenset({marriage.spouse_a_id, marriage.spouse_b_id}),
            marriage.married_at,
        )
        existing_marriage_by_key.setdefault(marriage_key, marriage)

    file_marriage_key_to_ref: dict[MarriageFileKey, str] = {}
    for marriage_row in parsed.marriages:
        file_key = (
            frozenset({marriage_row.spouse_a_ref, marriage_row.spouse_b_ref}),
            marriage_row.married_at,
        )
        if file_key in file_marriage_key_to_ref:
            match.marriage_duplicate_of[marriage_row.ref] = file_marriage_key_to_ref[
                file_key
            ]
        else:
            file_marriage_key_to_ref[file_key] = marriage_row.ref

        id_a = match.person_existing_id.get(marriage_row.spouse_a_ref)
        id_b = match.person_existing_id.get(marriage_row.spouse_b_ref)
        if id_a is None or id_b is None:
            continue
        existing_marriage = existing_marriage_by_key.get(
            (frozenset({id_a, id_b}), marriage_row.married_at)
        )
        if existing_marriage is not None:
            match.marriage_existing_id[marriage_row.ref] = existing_marriage.safe_id

    return match
